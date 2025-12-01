import os
import sys
import re
import csv
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime

if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)
else:
    base_path = os.path.dirname(__file__)

# Load Existing Buku Daftar excelbook
filename = os.path.join(base_path, "Buku_Daftar_UAT.xlsx") # Main Excel Master record
admin_path = os.path.join(base_path, "admin_password.csv") # Admin password for record edit
workbook = load_workbook(filename)
sheet = workbook.active

# Global variables
otp_verified = False # Start dictionary for password match
selected_row_index = None
invalid_attempts = 0 # invalid attempt counter for admin password trigger

# Find the existing table in the sheet
def get_table_range(sheet):
    for table in sheet.tables.values():
        return table.ref  # Returns the table range

# Find header of the column
column_indexes = {}
for col in sheet.iter_cols(1, sheet.max_column):
    header = col[0].value
    if header in ["REPORT NUMBER", "INTERNAL REFERENCE NUMBER", "DATE RECEIVED","RECEIVED BY","CONTACT PERSON","APPLICANT BY", "CLIENT",
            "WORK TITLE", "QUANTITY", "SAMPLE MARKING", "WORK CLASS", "TOT", "NO. OF TEST",
            "START TEST DATE", "END TEST DATE", "APPROVED DATE", "REPORT RELEASE DATE",
            "LABORATORY PERSONNEL","REVENUE/REMARKS"]:
        column_indexes[header] = col[0].column  # Store the column index

# Get admin password from csv for record editing purposes
def get_values_from_csv(admin_password):
    values = []
    with open(admin_password, mode='r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for index, row in enumerate(reader):
            if index == 0:  # First row
                values.append(row[1])  # admin password
                break  # Exit after getting the second row
    return values  # Return a list of values

# Read admin password CSV path
admin_password = admin_path

# Get the variable values (admin password) from the CSV
variable_values = get_values_from_csv(admin_password)

# Password value
admin_var = variable_values[0]

# Function to change the admin password
def change_admin_password():
    def save_new_password():
        new_password = entry_new_password.get()
        if not new_password:
            messagebox.showerror("Error", "Password cannot be empty.")
            return

        # Update the CSV file with the new password
        with open("admin_password.csv", "r") as f:
            rows = list(csv.reader(f))

        rows[0][1] = new_password  # Update the password in the second column of the first row

        with open("admin_password.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        # Program need to be exit-ed for new password to take effect
        messagebox.showinfo("Success", "Password updated successfully!. Exit program and try again")
        password_window.destroy()

    # Password tkinter window
    password_window = tk.Toplevel()
    password_window.title("Change Admin Password")
    password_window.geometry("300x150")

    tk.Label(password_window, text="Enter New Password:").pack(pady=10)
    entry_new_password = tk.Entry(password_window, show="*")
    entry_new_password.pack(pady=5)

    tk.Button(password_window, text="Save", command=save_new_password).pack(pady=10)

# Main window for reference number searching program
def main_window():
    global invalid_attempts # invalid attempt counter for admin password

    # Main window, main window 😂
    root = tk.Tk()
    root.title("Search Reference Number")
    root.geometry("600x150")

    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10)

    # Label for code so my program looks cool.
    #...Sorry I had too much coffee
    code_label = tk.Label(frame, text="PA/UAT").grid(row=1, column=0, padx=5, pady=5, sticky="w")

    # Work Class
    tk.Label(frame, text="Work Class").grid(row=0, column=1, padx=5, pady=5, sticky="w")

    workclass_var = tk.StringVar()  # Variable to store selected Work Class
    workclass_dropdown = ttk.Combobox(frame, textvariable=workclass_var, state="readonly")
    workclass_dropdown["values"] = ["9230", "9240", "9250", "9260"]  # Work Class options
    workclass_dropdown.grid(row=1, column=1, padx=5, pady=5)

    # Get the last two digits of the current year
    current_year_suffix = datetime.now().year % 100

    # well...aesthetics again
    year_label = tk.Label(frame, text=f'{current_year_suffix}').grid(row=1, column=2, padx=5, pady=5, sticky="w")

    tk.Label(frame, text="Running Number").grid(row=0, column=3, padx=5, pady=5, sticky="w")
    entry_running_number = tk.Entry(frame)
    entry_running_number.grid(row=1, column=3, padx=5, pady=5)

    # search part of the program
    def search_reference():
        global invalid_attempts
        workclass = workclass_var.get() # get workclass var from tkinter window
        running_number = entry_running_number.get() # get running number var from tkinter window

        if not workclass or not running_number:
            invalid_attempts += 1 # add one for each error attempt
            messagebox.showerror("Error", "Please enter a Reference Number.")

            # Trigger event after 3 invalid attempts
            if invalid_attempts == 3:
                invalid_attempts = 0  # Reset the counter
                on_no_reference() # triggering event for change password
            return

        # If valid, reset the counter
        invalid_attempts = 0

        reference_number = f"PA/UAT/{workclass}/{current_year_suffix}/{running_number}" # var for ref number

        # Locate the column index for "INTERNAL REFERENCE NUMBER" from master record
        col_index = column_indexes.get("INTERNAL REFERENCE NUMBER")
        if not col_index:
            messagebox.showerror("Error", "Column 'INTERNAL REFERENCE NUMBER' not found.")
            return

        # Search for the reference number in the column
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
            cell = row[0]
            cell_value = cell.value
            if cell.value:
                normalized_cell_value = re.split(r"-\(\d+\)$", cell_value)[0] # Algorithm to ignore suffix. $ looks at the end (\d+\) look at one or more digit at the end after the "-"
                if normalized_cell_value == reference_number:
                    row_data = cell.row
                    messagebox.showinfo("Found", f"Reference Number found at row {row_data}.")
                    root.destroy()  # Close the main window
                    show_options_window(reference_number, row_data)
                    return  # Exit the function if a match is found

        # If not found
        messagebox.showinfo("Not Found", "Reference Number not found in the table.")

        # Show options window if the reference number is found
        root.destroy()  # Close the main window
        show_options_window(reference_number, row_data)

    # Triggering event to change the password
    def on_no_reference():
        response = messagebox.askyesno("Change Admin Password",
                                       "Too many invalid attempts. Do you want to change the admin password?")
        if response:
            change_admin_password()

    # Back to the main window. This is the search button for the main window
    tk.Button(frame, text="Search", command=search_reference).grid(row=1, column=5, columnspan=2, pady=10)
    root.mainloop()

# Once main window is passed,you get two option. This is that option tkint window.
# We pass reference_number and row_data in this
def show_options_window(reference_number, row_data):
    options_window = tk.Tk()
    options_window.title("Options")
    options_window.geometry("400x200")

    frame = tk.Frame(options_window)
    frame.pack(padx=10, pady=10)

    tk.Label(frame, text=f"Reference Number: {reference_number}", font=("Arial", 12)).pack(pady=10)

    # Option 1, kemaskini window to amend date of test and revenue
    def open_kemaskini():
        options_window.destroy()
        kemaskini(reference_number, row_data)

    # Option 2, show record and edit if needed "wink2"
    def open_papar_rekod():
        options_window.destroy()
        papar_rekod(reference_number, row_data)

    tk.Button(frame, text="Kemaskini", command=open_kemaskini, width=15).pack(pady=5)
    tk.Button(frame, text="Papar Rekod", command=open_papar_rekod, width=15).pack(pady=5)

    options_window.mainloop()

# Option 1 window. Reference_number is not needed anymore but I am too coward to delete it
def kemaskini(reference_number, row_data):
    new_data = tk.Tk()
    new_data.title("Kemaskini Data Pengujian")
    new_data.geometry("800x600")

    frame = tk.Frame(new_data)
    frame.pack(padx=10, pady=10)

    # Start Test Date
    tk.Label(frame, text="Start Test Date").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_SD = tk.Entry(frame)
    entry_SD.grid(row=0, column=1, padx=5, pady=5)

    # End Test Date
    tk.Label(frame, text="End Test Date").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    entry_ED = tk.Entry(frame)
    entry_ED.grid(row=1, column=1, padx=5, pady=5)

    # Approved Date
    tk.Label(frame, text="Approved Date").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    entry_AD = tk.Entry(frame)
    entry_AD.grid(row=2, column=1, padx=5, pady=5)

    # Release Date
    tk.Label(frame, text="Report Release Date").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    entry_RD = tk.Entry(frame)
    entry_RD.grid(row=3, column=1, padx=5, pady=5)

    # Laboratory Personnel
    tk.Label(frame, text="Laboratory Personnel").grid(row=4, column=0, padx=5, pady=5, sticky="e")
    entry_LP = tk.Entry(frame)
    entry_LP.grid(row=4, column=1, padx=5, pady=5)

    # Remarks/Revenue
    tk.Label(frame, text="Revenue/Remarks").grid(row=5, column=0, padx=5, pady=5, sticky="e")
    entry_RR = tk.Entry(frame)
    entry_RR.grid(row=5, column=1, padx=5, pady=5)

    def save_updates():
        start_test_date = entry_SD.get()
        end_test_date = entry_ED.get()
        approved_date = entry_AD.get()
        release_date = entry_RD.get()
        lab_personnel = entry_LP.get()
        revenue_remarks = entry_RR.get()

        if "START TEST DATE" in column_indexes:
            sheet.cell(row=row_data, column=column_indexes["START TEST DATE"], value=start_test_date)
        if "END TEST DATE" in column_indexes:
            sheet.cell(row=row_data, column=column_indexes["END TEST DATE"], value=end_test_date)
        if "APPROVED DATE" in column_indexes:
            sheet.cell(row=row_data, column=column_indexes["APPROVED DATE"], value=approved_date)
        if "REPORT RELEASE DATE" in column_indexes:
            sheet.cell(row=row_data, column=column_indexes["REPORT RELEASE DATE"], value=release_date)
        if "LABORATORY PERSONNEL" in column_indexes:
            sheet.cell(row=row_data, column=column_indexes["LABORATORY PERSONNEL"], value=lab_personnel)
        if "REVENUE/REMARKS" in column_indexes:
            sheet.cell(row=row_data, column=column_indexes["REVENUE/REMARKS"], value=revenue_remarks)

        # Save the workbook
        workbook.save(filename)
        messagebox.showinfo("Success", "Data updated successfully.")
        new_data.destroy()

    # Save Button. Yes, I know, I didn't include the error if empty field
    # ...and also what would happen if someone clicked Save without entering any data
    tk.Button(frame, text="Save", command=save_updates).grid(row=6, column=0, columnspan=2, pady=20)

    # Here you go, Are you happy now !!...gotta stop chugging coffee
    tk.Label(frame, text="Jangan biarkan medan kosong sebelum menekan butang Save", font=("Arial", 12), fg="red").grid(row=7, column=0, columnspan=2, pady=20)

    new_data.mainloop()

# Option 2. Yes I know, the reference number isn't needed anymore
def papar_rekod(reference_number, row_data):
    def load_row_data(row_data):
        # Extract all the row value based on row number of the RN "row_data"
        row_values = {}
        for cell in sheet[row_data]:
            column_header = sheet.cell(row=1, column=cell.column).value  # Assuming headers are in row 1
            row_values[column_header] = cell.value
        return row_values

    # Function to save papar data. Give error if no password
    def save_papar_data():
        global sheet
        if not otp_verified:
            messagebox.showerror("Error", "You must Logged in as admin before saving changes.")
            return

        # Data to update. Extract from papar main window
        updated_data = {
            "REPORT NUMBER": entry_RN.get(),
            "INTERNAL REFERENCE NUMBER": entry_IRN.get(),
            "CLIENT": entry_client.get(),
            "WORK TITLE": entry_worktitle.get(),
            "DATE RECEIVED": entry_date_received.get(),
            "CONTACT PERSON":entry_CP.get(),
            "RECEIVED BY": entry_Rby.get(),
            "QUANTITY": entry_kuantiti.get(),
            "WORK CLASS": entry_workclass.get(),
            "SAMPLE MARKING": entry_sample_marking.get(),
        }

        # Loop through the updated data and save it to the corresponding column in the sheet
        for header, value in updated_data.items():
            if header in column_indexes:
                column_index = column_indexes[header]
                sheet.cell(row=row_data, column=column_index, value=value)

        try:
            # Save the updated data back to the Excel file
            workbook.save(filename)
            messagebox.showinfo("Success", "Data saved successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Error saving data: {e}")

    # Prompt for admin login. Main prompt #1
    def prompt_otp(papar_rekod):
        global otp_popup, otp_entry
        otp_popup = tk.Toplevel(papar_rekod)
        otp_popup.title("Admin Login")
        otp_popup.geometry("300x300")

        tk.Label(otp_popup, text="Enter Password:").pack(pady=10)
        otp_entry = tk.Entry(otp_popup, show="*")
        otp_entry.pack(pady=10)
        tk.Button(otp_popup, text="Submit", command=verify_otp).pack(pady=10)

    def verify_otp():
        global otp_verified
        if otp_entry.get() == admin_var:  # Password from csv
            otp_verified = True
            otp_popup.destroy()
            messagebox.showinfo("Success", "Password verified! You can now edit the fields.")
        else:
            messagebox.showerror("Error", "Invalid Password! Try again.")

    # Restrict field editing initially. If field click trigger prompt_otp i.e the main prompt #1
    def on_field_click(event):
        if not otp_verified:
            prompt_otp(papar)

    # Regurgitate row values. To be populated inside respective tkinter entry
    row_values = load_row_data(row_data)

    # Main papar window for our option 2
    papar = tk.Tk()
    papar.title("Rekod Makmal")
    papar.geometry("800x600")

    frame = tk.Frame(papar)
    frame.pack(padx=10, pady=10)

    # Report Number
    tk.Label(frame, text="Report Number").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_RN = tk.Entry(frame)
    entry_RN.grid(row=0, column=1, padx=5, pady=5)
    entry_RN.insert(0, row_values.get("REPORT NUMBER", "")) # populate entry field with data from excel

    # Internal Reference Number
    tk.Label(frame, text="Internal Reference Number").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    entry_IRN = tk.Entry(frame)
    entry_IRN.grid(row=1, column=1, padx=5, pady=5)
    entry_IRN.insert(0, row_values.get("INTERNAL REFERENCE NUMBER", "")) # Baca kat atas cukuplah jenuh la nak ulang

    # Client
    tk.Label(frame, text="Client").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    entry_client = tk.Entry(frame)
    entry_client.grid(row=2, column=1, padx=5, pady=5)
    entry_client.insert(0, row_values.get("CLIENT", "")) # Harey baca kat atas la !

    # Contact Person
    tk.Label(frame, text="Contact Person").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    entry_CP = tk.Entry(frame)
    entry_CP.grid(row=3, column=1, padx=5, pady=5)
    entry_CP.insert(0, row_values.get("CONTACT PERSON", "")) # @#$%*& !!!

    # Work Title
    tk.Label(frame, text="Work Title").grid(row=4, column=0, padx=5, pady=5, sticky="e")
    entry_worktitle = tk.Entry(frame)
    entry_worktitle.grid(row=4, column=1, padx=5, pady=5)
    entry_worktitle.insert(0, row_values.get("WORK TITLE", ""))

    # Date Received
    tk.Label(frame, text="Date Received").grid(row=5, column=0, padx=5, pady=5, sticky="e")
    entry_date_received = tk.Entry(frame)
    entry_date_received.grid(row=5, column=1, padx=5, pady=5)
    entry_date_received.insert(0, row_values.get("DATE RECEIVED", ""))

    # Received by
    tk.Label(frame, text="Received By").grid(row=6, column=0, padx=5, pady=5, sticky="e")
    entry_Rby = tk.Entry(frame)
    entry_Rby.grid(row=6, column=1, padx=5, pady=5)
    entry_Rby.insert(0, row_values.get("RECEIVED BY", ""))

    # Quantity
    tk.Label(frame, text="Kuantiti").grid(row=7, column=0, padx=5, pady=5, sticky="e")
    entry_kuantiti = tk.Entry(frame)
    entry_kuantiti.grid(row=7, column=1, padx=5, pady=5)
    entry_kuantiti.insert(0, row_values.get("QUANTITY", ""))

    # Workclass
    tk.Label(frame, text="Workclass").grid(row=8, column=0, padx=5, pady=5, sticky="e")
    entry_workclass = tk.Entry(frame)
    entry_workclass.grid(row=8, column=1, padx=5, pady=5)
    entry_workclass.insert(0, row_values.get("WORK CLASS", ""))

    # Sample Marking
    tk.Label(frame, text="Sample Marking").grid(row=9, column=0, padx=5, pady=5, sticky="e")
    entry_sample_marking = tk.Entry(frame)
    entry_sample_marking.grid(row=9, column=1, padx=5, pady=5)
    entry_sample_marking.insert(0, row_values.get("SAMPLE MARKING", ""))

    # Groupkan entries so that we could create a triggering event for our OTP...password admin
    entries = [entry_RN, entry_IRN, entry_client, entry_CP, entry_worktitle, entry_date_received,
               entry_Rby, entry_kuantiti, entry_workclass, entry_sample_marking]

    for entry in entries:
        entry.bind("<FocusIn>", on_field_click)

    # Save button (optional functionality)
    tk.Button(frame, text="Save Data", command=save_papar_data).grid(row=10, columnspan=3, pady=10)
    papar.mainloop()

if __name__ == "__main__":
    main_window()

